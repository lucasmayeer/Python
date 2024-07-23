import openpyxl
import requests

# Função para consultar dados do CNPJ na Receita Federal
def consultar_cnpj(cnpj):
    url = f"https://brasilapi.com.br/api/cnpj/v1/{cnpj}"
    response = requests.get(url)
    if response.status_code == 200:
        return response.json()
    else:
        return None

# Função para salvar dados em um novo arquivo Excel
def salvar_dados_em_arquivo(cnpj, dados):
    nome_arquivo = f"C:/Users/Lucas/OneDrive/Desktop/VOX/dados_cnpj_{cnpj}.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Dados CNPJ"
    # Adicionando os dados em colunas separadas
    for row_index, item in enumerate(dados.items(), start=1):
        key, value = item
        sheet.cell(row=row_index, column=1, value=key)
        sheet.cell(row=row_index, column=2, value=value)
    workbook.save(nome_arquivo)

# Carregar CNPJs da planilha Excel
def carregar_cnpjs_da_planilha(nome_arquivo):
    cnpjs = []
    try:
        workbook = openpyxl.load_workbook(nome_arquivo)
        sheet = workbook.active
        for cell in sheet["B"]:
            if cell.value is not None:
                cnpjs.append(cell.value)
            else:
                break
    except FileNotFoundError:
        print(f"O arquivo {nome_arquivo} não foi encontrado.")
    return cnpjs

# Nome do arquivo Excel
nome_arquivo = "C:/Users/Lucas/OneDrive/Desktop/VOX/contatos_python.xlsx"

# Carregar CNPJs da planilha
cnpjs = carregar_cnpjs_da_planilha(nome_arquivo)

# Iterar sobre os CNPJs e consultar dados
for cnpj in cnpjs:
    dados = consultar_cnpj(cnpj)
    if dados:
        # Salvar dados em um arquivo separado
        salvar_dados_em_arquivo(cnpj, dados)
        print(f"Dados do CNPJ {cnpj} salvos com sucesso!")
    else:
        print(f"Erro ao consultar CNPJ {cnpj}")
